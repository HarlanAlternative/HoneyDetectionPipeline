#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PowerBI Integration for Industrial Honey Quality Pipeline
Provides data model definition, template generation and data export
"""

import os
import json
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
import warnings

# Database connectivity
try:
    from sqlalchemy import create_engine, text
    from sqlalchemy.pool import QueuePool
    DB_LIBS_AVAILABLE = True
except ImportError:
    DB_LIBS_AVAILABLE = False
    warnings.warn("Database libraries not available")

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_LIBS_AVAILABLE = True
except ImportError:
    EXCEL_LIBS_AVAILABLE = False
    warnings.warn("Excel export libraries not available")


@dataclass
class PowerBIConfig:
    """Configuration for PowerBI integration"""
    enabled: bool
    refresh_schedule: str
    data_sources: List[str]
    export_formats: List[str]
    template_directory: str


class PowerBIIntegration:
    """PowerBI integration for data export and template generation"""
    
    def __init__(self, config_path: str = "config/etl_config.json"):
        """Initialize PowerBI integration"""
        self.config = self._load_config(config_path)
        self.db_engine = None
        self.template_dir = self.config.template_directory
        
        # Ensure template directory exists
        os.makedirs(self.template_dir, exist_ok=True)
        
        if DB_LIBS_AVAILABLE:
            self.db_engine = self._init_database()
    
    def _load_config(self, config_path: str) -> PowerBIConfig:
        """Load PowerBI configuration"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            
            powerbi_config = config_data.get('powerbi', {})
            
            return PowerBIConfig(
                enabled=powerbi_config.get('enabled', True),
                refresh_schedule=powerbi_config.get('refresh_schedule', '0 */6 * * *'),
                data_sources=powerbi_config.get('data_sources', []),
                export_formats=powerbi_config.get('export_formats', ['csv', 'xlsx', 'json']),
                template_directory=powerbi_config.get('template_directory', 'powerbi_templates')
            )
        except Exception as e:
            # Default configuration
            return PowerBIConfig(
                enabled=True,
                refresh_schedule='0 */6 * * *',
                data_sources=['honey_quality_data', 'quality_metrics', 'pipeline_status'],
                export_formats=['csv', 'xlsx', 'json'],
                template_directory='powerbi_templates'
            )
    
    def _init_database(self) -> Optional[Any]:
        """Initialize database connection"""
        try:
            # This would typically connect to the configured database
            # For now, return None as we're working with sample data
            return None
        except Exception as e:
            warnings.warn(f"Database connection failed: {e}")
            return None
    
    def create_powerbi_data_model(self) -> Dict[str, Any]:
        """Create PowerBI data model definition"""
        data_model = {
            "name": "Honey Quality Analysis",
            "description": "Data model for honey quality analysis and reporting",
            "version": "1.0.0",
            "created_at": datetime.now().isoformat(),
            "tables": [
                {
                    "name": "honey_quality_data",
                    "description": "Main honey quality measurements",
                    "columns": [
                        {"name": "id", "type": "int", "description": "Unique identifier"},
                        {"name": "batch_id", "type": "string", "description": "Batch identifier"},
                        {"name": "sample_id", "type": "string", "description": "Sample identifier"},
                        {"name": "moisture", "type": "decimal", "description": "Moisture content (%)"},
                        {"name": "ph", "type": "decimal", "description": "pH value"},
                        {"name": "diastase_activity", "type": "decimal", "description": "Diastase activity"},
                        {"name": "h_m_f", "type": "decimal", "description": "Hydroxymethylfurfural content"},
                        {"name": "quality_score", "type": "decimal", "description": "Calculated quality score"},
                        {"name": "quality_category", "type": "string", "description": "Quality category"},
                        {"name": "collection_date", "type": "date", "description": "Sample collection date"},
                        {"name": "lab_id", "type": "string", "description": "Laboratory identifier"},
                        {"name": "analyst", "type": "string", "description": "Analyst name"},
                        {"name": "processed_at", "type": "datetime", "description": "Processing timestamp"}
                    ]
                },
                {
                    "name": "quality_metrics",
                    "description": "Aggregated quality metrics",
                    "columns": [
                        {"name": "date", "type": "date", "description": "Date"},
                        {"name": "total_samples", "type": "int", "description": "Total samples processed"},
                        {"name": "avg_quality_score", "type": "decimal", "description": "Average quality score"},
                        {"name": "excellent_count", "type": "int", "description": "Excellent quality samples"},
                        {"name": "good_count", "type": "int", "description": "Good quality samples"},
                        {"name": "fair_count", "type": "int", "description": "Fair quality samples"},
                        {"name": "poor_count", "type": "int", "description": "Poor quality samples"}
                    ]
                },
                {
                    "name": "pipeline_status",
                    "description": "ETL pipeline execution status",
                    "columns": [
                        {"name": "pipeline_name", "type": "string", "description": "Pipeline identifier"},
                        {"name": "status", "type": "string", "description": "Execution status"},
                        {"name": "start_time", "type": "datetime", "description": "Start timestamp"},
                        {"name": "end_time", "type": "datetime", "description": "End timestamp"},
                        {"name": "duration_seconds", "type": "decimal", "description": "Execution duration"},
                        {"name": "records_processed", "type": "int", "description": "Records processed"},
                        {"name": "error_message", "type": "string", "description": "Error details if failed"}
                    ]
                }
            ],
            "relationships": [
                {
                    "name": "quality_data_to_metrics",
                    "from_table": "honey_quality_data",
                    "from_column": "collection_date",
                    "to_table": "quality_metrics",
                    "to_column": "date",
                    "type": "many_to_one"
                }
            ]
        }
        
        return data_model
    
    def generate_powerbi_template(self) -> str:
        """Generate PowerBI template files"""
        try:
            # Create data model
            data_model = self.create_powerbi_data_model()
            
            # Save data model
            model_file = os.path.join(self.template_dir, "data_model.json")
            with open(model_file, 'w', encoding='utf-8') as f:
                json.dump(data_model, f, indent=2, ensure_ascii=False, default=str)
            
            # Create connection string template
            connection_template = self._create_connection_template()
            conn_file = os.path.join(self.template_dir, "connection_string.txt")
            with open(conn_file, 'w', encoding='utf-8') as f:
                f.write(connection_template)
            
            # Create sample queries
            queries_template = self._create_sample_queries()
            queries_file = os.path.join(self.template_dir, "sample_queries.sql")
            with open(queries_file, 'w', encoding='utf-8') as f:
                f.write(queries_template)
            
            # Create report design guide
            design_guide = self._create_report_design_guide()
            guide_file = os.path.join(self.template_dir, "report_design_guide.md")
            with open(guide_file, 'w', encoding='utf-8') as f:
                f.write(design_guide)
            
            print(f"PowerBI templates generated successfully in: {self.template_dir}")
            return self.template_dir
            
        except Exception as e:
            print(f"PowerBI template generation failed: {e}")
            raise
    
    def _create_connection_template(self) -> str:
        """Create database connection string template"""
        return """# Database Connection String Template

# PostgreSQL Connection
Server=localhost;Database=honey_warehouse;Port=5432;Uid=postgres;Pwd=password;

# SQL Server Connection (if using SQL Server)
Server=localhost;Database=honey_warehouse;Trusted_Connection=true;

# MySQL Connection (if using MySQL)
Server=localhost;Database=honey_warehouse;Uid=root;Pwd=password;

# Connection Parameters
- Server: Database server address
- Database: Database name (honey_warehouse)
- Port: Database port (default: 5432 for PostgreSQL)
- Uid: Username
- Pwd: Password
- Trusted_Connection: Use Windows authentication (SQL Server)

# Notes
- Update server address and credentials as needed
- Ensure database user has read permissions
- Test connection before using in PowerBI
"""
    
    def _create_sample_queries(self) -> str:
        """Create sample SQL queries for PowerBI"""
        return """-- Sample SQL Queries for PowerBI Integration

-- 1. Basic Quality Data Query
SELECT 
    batch_id,
    sample_id,
    moisture,
    ph,
    diastase_activity,
    h_m_f,
    quality_score,
    quality_category,
    collection_date,
    lab_id,
    analyst
FROM honey_quality_data
WHERE collection_date >= CURRENT_DATE - INTERVAL '30 days'
ORDER BY collection_date DESC;

-- 2. Quality Metrics Summary
SELECT 
    DATE(collection_date) as date,
    COUNT(*) as total_samples,
    AVG(quality_score) as avg_quality_score,
    COUNT(CASE WHEN quality_category = 'Excellent' THEN 1 END) as excellent_count,
    COUNT(CASE WHEN quality_category = 'Good' THEN 1 END) as good_count,
    COUNT(CASE WHEN quality_category = 'Fair' THEN 1 END) as fair_count,
    COUNT(CASE WHEN quality_category = 'Poor' THEN 1 END) as poor_count
FROM honey_quality_data
GROUP BY DATE(collection_date)
ORDER BY date DESC;

-- 3. Laboratory Performance Analysis
SELECT 
    lab_id,
    COUNT(*) as total_samples,
    AVG(quality_score) as avg_quality_score,
    COUNT(CASE WHEN quality_category = 'Excellent' THEN 1 END) as excellent_count,
    ROUND(
        COUNT(CASE WHEN quality_category = 'Excellent' THEN 1 END) * 100.0 / COUNT(*), 2
    ) as excellence_rate
FROM honey_quality_data
GROUP BY lab_id
ORDER BY excellence_rate DESC;

-- 4. Quality Trend Analysis
SELECT 
    DATE_TRUNC('week', collection_date) as week_start,
    COUNT(*) as samples_per_week,
    AVG(quality_score) as weekly_avg_score,
    AVG(moisture) as weekly_avg_moisture,
    AVG(ph) as weekly_avg_ph
FROM honey_quality_data
WHERE collection_date >= CURRENT_DATE - INTERVAL '12 weeks'
GROUP BY DATE_TRUNC('week', collection_date)
ORDER BY week_start;

-- 5. Parameter Distribution Analysis
SELECT 
    'Moisture' as parameter,
    MIN(moisture) as min_value,
    MAX(moisture) as max_value,
    AVG(moisture) as avg_value,
    STDDEV(moisture) as std_dev
FROM honey_quality_data
WHERE collection_date >= CURRENT_DATE - INTERVAL '30 days'

UNION ALL

SELECT 
    'pH' as parameter,
    MIN(ph) as min_value,
    MAX(ph) as max_value,
    AVG(ph) as avg_value,
    STDDEV(ph) as std_dev
FROM honey_quality_data
WHERE collection_date >= CURRENT_DATE - INTERVAL '30 days'

UNION ALL

SELECT 
    'Diastase Activity' as parameter,
    MIN(diastase_activity) as min_value,
    MAX(diastase_activity) as max_value,
    AVG(diastase_activity) as avg_value,
    STDDEV(diastase_activity) as std_dev
FROM honey_quality_data
WHERE collection_date >= CURRENT_DATE - INTERVAL '30 days'

UNION ALL

SELECT 
    'HMF' as parameter,
    MIN(h_m_f) as min_value,
    MAX(h_m_f) as max_value,
    AVG(h_m_f) as avg_value,
    STDDEV(h_m_f) as std_dev
FROM honey_quality_data
WHERE collection_date >= CURRENT_DATE - INTERVAL '30 days';
"""
    
    def _create_report_design_guide(self) -> str:
        """Create PowerBI report design guide"""
        return """# PowerBI Report Design Guide for Honey Quality Analysis

## Overview
This guide provides recommendations for creating effective PowerBI reports for honey quality analysis.

## Recommended Visualizations

### 1. Quality Dashboard
- **KPI Cards**: Total samples, average quality score, compliance rate
- **Gauge Charts**: Quality score distribution, parameter compliance
- **Bar Charts**: Quality category distribution, laboratory performance
- **Line Charts**: Quality trends over time

### 2. Parameter Analysis
- **Scatter Plots**: Moisture vs pH, Diastase vs HMF
- **Histograms**: Parameter distribution analysis
- **Box Plots**: Parameter range and outliers
- **Heat Maps**: Correlation between parameters

### 3. Time Series Analysis
- **Line Charts**: Quality trends, parameter trends
- **Area Charts**: Cumulative quality metrics
- **Waterfall Charts**: Quality improvement tracking

## Color Scheme Recommendations
- **Excellent Quality**: Green (#00FF00)
- **Good Quality**: Light Green (#90EE90)
- **Fair Quality**: Yellow (#FFFF00)
- **Poor Quality**: Red (#FF0000)
- **Neutral**: Gray (#808080)

## Filtering Strategy
- **Date Range**: Last 7, 30, 90 days
- **Laboratory**: Individual lab selection
- **Quality Category**: Filter by quality level
- **Parameter Ranges**: Filter by parameter values

## Refresh Strategy
- **Real-time**: Every 6 hours (configurable)
- **Manual**: On-demand refresh
- **Scheduled**: Daily at 6 AM

## Data Source Configuration
1. Connect to database using provided connection string
2. Import required tables
3. Set up relationships between tables
4. Configure data refresh schedule
5. Test data connectivity

## Best Practices
- Use consistent naming conventions
- Implement proper error handling
- Optimize query performance
- Regular data validation
- User access control
- Backup and recovery procedures

## Troubleshooting
- Verify database connectivity
- Check data refresh permissions
- Monitor query performance
- Validate data quality
- Review error logs
"""
    
    def create_sample_data_for_powerbi(self, num_records: int = 1000) -> pd.DataFrame:
        """Create sample data for PowerBI testing"""
        np.random.seed(42)
        
        # Generate sample data
        data = {
            'batch_id': [f'BATCH_{i:03d}' for i in range(1, num_records + 1)],
            'sample_id': [f'SAMPLE_{i:06d}' for i in range(1, num_records + 1)],
            'moisture': np.random.uniform(15.0, 20.0, num_records),
            'ph': np.random.uniform(3.5, 6.5, num_records),
            'diastase_activity': np.random.uniform(8.0, 15.0, num_records),
            'h_m_f': np.random.uniform(20.0, 40.0, num_records),
            'collection_date': pd.date_range('2024-01-01', periods=num_records, freq='H'),
            'lab_id': np.random.choice(['LAB_A', 'LAB_B', 'LAB_C', 'LAB_D'], num_records),
            'analyst': np.random.choice(['Analyst_1', 'Analyst_2', 'Analyst_3', 'Analyst_4'], num_records)
        }
        
        df = pd.DataFrame(data)
        
        # Calculate quality score
        df['quality_score'] = df.apply(self._calculate_sample_quality_score, axis=1)
        df['quality_category'] = df['quality_score'].apply(self._categorize_sample_quality)
        
        # Add processing timestamp
        df['processed_at'] = datetime.now()
        
        return df
    
    def _calculate_sample_quality_score(self, row: pd.Series) -> float:
        """Calculate quality score for sample data"""
        score = 100.0
        
        # Moisture scoring
        if 15.0 <= row['moisture'] <= 20.0:
            score -= 0
        elif 14.0 <= row['moisture'] <= 21.0:
            score -= 5
        else:
            score -= 15
        
        # pH scoring
        if 3.5 <= row['ph'] <= 6.5:
            score -= 0
        elif 3.0 <= row['ph'] <= 7.0:
            score -= 5
        else:
            score -= 15
        
        # Diastase activity scoring
        if row['diastase_activity'] >= 8.0:
            score -= 0
        elif row['diastase_activity'] >= 6.0:
            score -= 5
        else:
            score -= 15
        
        # HMF scoring
        if row['h_m_f'] <= 40.0:
            score -= 0
        elif row['h_m_f'] <= 50.0:
            score -= 5
        else:
            score -= 15
        
        return max(0.0, score)
    
    def _categorize_sample_quality(self, score: float) -> str:
        """Categorize quality based on score"""
        if score >= 90:
            return "Excellent"
        elif score >= 80:
            return "Good"
        elif score >= 70:
            return "Fair"
        else:
            return "Poor"
    
    def export_to_powerbi_format(self, data: pd.DataFrame, format_type: str = "xlsx") -> str:
        """Export data to PowerBI-compatible format"""
        try:
            output_dir = "powerbi_data"
            os.makedirs(output_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if format_type == "xlsx":
                filename = f"honey_quality_data_{timestamp}.xlsx"
                filepath = os.path.join(output_dir, filename)
                
                if EXCEL_LIBS_AVAILABLE:
                    self._export_to_excel(data, filepath)
                else:
                    # Fallback to CSV if Excel not available
                    filename = f"honey_quality_data_{timestamp}.csv"
                    filepath = os.path.join(output_dir, filename)
                    data.to_csv(filepath, index=False)
                    
            elif format_type == "csv":
                filename = f"honey_quality_data_{timestamp}.csv"
                filepath = os.path.join(output_dir, filename)
                data.to_csv(filepath, index=False)
                
            elif format_type == "json":
                filename = f"honey_quality_data_{timestamp}.json"
                filepath = os.path.join(output_dir, filename)
                data.to_json(filepath, orient='records', indent=2, date_format='iso')
                
            else:
                raise ValueError(f"Unsupported export format: {format_type}")
            
            print(f"Data exported to PowerBI format: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"PowerBI export failed: {e}")
            raise
    
    def _export_to_excel(self, data: pd.DataFrame, filepath: str):
        """Export data to Excel with formatting"""
        if not EXCEL_LIBS_AVAILABLE:
            raise ImportError("Excel export libraries not available")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Honey Quality Data"
        
        # Add data
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Format header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        wb.close()
    
    def cleanup(self):
        """Cleanup resources"""
        try:
            if self.db_engine:
                self.db_engine.dispose()
            
            print("PowerBI integration cleaned up successfully")
            
        except Exception as e:
            print(f"PowerBI cleanup failed: {e}")


if __name__ == "__main__":
    # Example usage
    pbi = PowerBIIntegration()
    
    try:
        # Generate templates
        pbi.generate_powerbi_template()
        
        # Create sample data
        sample_data = pbi.create_sample_data_for_powerbi(100)
        
        # Export to PowerBI format
        export_path = pbi.export_to_powerbi_format(sample_data, "xlsx")
        print(f"Sample data exported to: {export_path}")
        
    finally:
        pbi.cleanup()
