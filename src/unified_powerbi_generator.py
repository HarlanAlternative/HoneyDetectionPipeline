#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Unified PowerBI Report Generator
Combines all PowerBI generation functionality into one module
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
import warnings

# Excel processing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.chart import BarChart, LineChart, PieChart
    EXCEL_LIBS_AVAILABLE = True
except ImportError:
    EXCEL_LIBS_AVAILABLE = False
    warnings.warn("Excel chart libraries not available")


class UnifiedPowerBIGenerator:
    """Unified PowerBI report generator with all functionality"""
    
    def __init__(self, output_dir: str = "powerbi_reports"):
        """Initialize the generator"""
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_sample_data(self, num_records: int = 1000) -> pd.DataFrame:
        """Generate comprehensive sample data for reporting"""
        np.random.seed(42)
        
        # Generate time series data
        start_date = datetime(2024, 1, 1)
        dates = pd.date_range(start_date, periods=num_records, freq='H')
        
        data = {
            'batch_id': [f'BATCH_{i:03d}' for i in range(1, num_records + 1)],
            'sample_id': [f'SAMPLE_{i:06d}' for i in range(1, num_records + 1)],
            'moisture': np.random.uniform(15.0, 20.0, num_records),
            'ph': np.random.uniform(3.5, 6.5, num_records),
            'diastase_activity': np.random.uniform(8.0, 15.0, num_records),
            'h_m_f': np.random.uniform(20.0, 40.0, num_records),
            'collection_date': dates,
            'lab_id': np.random.choice(['LAB_A', 'LAB_B', 'LAB_C', 'LAB_D'], num_records),
            'analyst': np.random.choice(['Analyst_1', 'Analyst_2', 'Analyst_3', 'Analyst_4'], num_records),
            'region': np.random.choice(['North', 'South', 'East', 'West'], num_records),
            'batch_size': np.random.choice([100, 200, 300, 500], num_records)
        }
        
        df = pd.DataFrame(data)
        
        # Calculate quality metrics
        df['quality_score'] = df.apply(self._calculate_quality_score, axis=1)
        df['quality_category'] = df['quality_score'].apply(self._categorize_quality)
        df['compliance_status'] = df['quality_score'].apply(self._get_compliance_status)
        
        # Add derived metrics for PowerBI
        df['moisture_deviation'] = abs(df['moisture'] - 17.5)
        df['ph_deviation'] = abs(df['ph'] - 5.0)
        df['processing_hour'] = df['collection_date'].dt.hour
        df['processing_day'] = df['collection_date'].dt.day_name()
        df['month'] = df['collection_date'].dt.month
        df['quarter'] = df['collection_date'].dt.quarter
        
        return df
    
    def _calculate_quality_score(self, row: pd.Series) -> float:
        """Calculate comprehensive quality score"""
        score = 100.0
        
        # Moisture scoring (target: 17.5%)
        moisture_dev = abs(row['moisture'] - 17.5)
        if moisture_dev <= 1.0:
            score -= 0
        elif moisture_dev <= 2.0:
            score -= 5
        elif moisture_dev <= 3.0:
            score -= 10
        else:
            score -= 20
        
        # pH scoring (target: 5.0)
        ph_dev = abs(row['ph'] - 5.0)
        if ph_dev <= 0.5:
            score -= 0
        elif ph_dev <= 1.0:
            score -= 5
        elif ph_dev <= 1.5:
            score -= 10
        else:
            score -= 20
        
        # Diastase activity scoring (min: 8.0)
        if row['diastase_activity'] >= 8.0:
            score -= 0
        elif row['diastase_activity'] >= 6.0:
            score -= 5
        else:
            score -= 15
        
        # HMF scoring (max: 40.0)
        if row['h_m_f'] <= 40.0:
            score -= 0
        elif row['h_m_f'] <= 50.0:
            score -= 5
        else:
            score -= 15
        
        return max(0.0, score)
    
    def _categorize_quality(self, score: float) -> str:
        """Categorize quality based on score"""
        if score >= 95:
            return "Premium"
        elif score >= 90:
            return "Excellent"
        elif score >= 80:
            return "Good"
        elif score >= 70:
            return "Fair"
        else:
            return "Poor"
    
    def _get_compliance_status(self, score: float) -> str:
        """Get compliance status"""
        if score >= 80:
            return "Compliant"
        elif score >= 70:
            return "Warning"
        else:
            return "Non-Compliant"
    
    def generate_excel_report(self, data: pd.DataFrame) -> str:
        """Generate comprehensive Excel report with charts"""
        if not EXCEL_LIBS_AVAILABLE:
            raise ImportError("Excel libraries not available")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"honey_quality_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quality Dashboard"
        
        # Title
        ws['A1'] = "Honey Quality Analysis - Dashboard"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # KPI Section
        kpi_data = [
            ["Total Samples", len(data), "A3"],
            ["Average Quality Score", f"{data['quality_score'].mean():.2f}", "A4"],
            ["Premium Quality Rate", f"{len(data[data['quality_category']=='Premium'])/len(data)*100:.1f}%", "A5"],
            ["Compliance Rate", f"{len(data[data['compliance_status']=='Compliant'])/len(data)*100:.1f}%", "A6"],
            ["Laboratories", data['lab_id'].nunique(), "A7"],
            ["Date Range", f"{data['collection_date'].min().strftime('%Y-%m-%d')} to {data['collection_date'].max().strftime('%Y-%m-%d')}", "A8"]
        ]
        
        for label, value, cell in kpi_data:
            ws[cell] = label
            ws[cell].font = Font(bold=True)
            ws[f"{chr(ord(cell[0])+1)}{cell[1:]}"] = value
        
        # Quality Distribution Chart
        quality_counts = data['quality_category'].value_counts()
        ws['A10'] = "Quality Distribution"
        ws['A10'].font = Font(bold=True, size=14)
        
        ws['A11'] = "Category"
        ws['B11'] = "Count"
        ws['A11'].font = Font(bold=True)
        ws['B11'].font = Font(bold=True)
        
        categories = ['Premium', 'Excellent', 'Good', 'Fair', 'Poor']
        for i, category in enumerate(categories):
            ws[f'A{12+i}'] = category
            ws[f'B{12+i}'] = quality_counts.get(category, 0)
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Quality Distribution"
        data_series = openpyxl.chart.Reference(ws, min_col=2, min_row=11, max_row=16)
        labels = openpyxl.chart.Reference(ws, min_col=1, min_row=12, max_row=16)
        chart.add_data(data_series, titles_from_data=False)
        chart.set_categories(labels)
        chart.height = 15
        chart.width = 20
        
        ws.add_chart(chart, "D10")
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        print(f"Excel report generated: {filepath}")
        return filepath
    
    def create_powerbi_template_files(self, data: pd.DataFrame) -> dict:
        """Create comprehensive PowerBI template files"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 1. Create enhanced data files
        csv_file = os.path.join(self.output_dir, f"honey_quality_data_{timestamp}.csv")
        excel_file = os.path.join(self.output_dir, f"honey_quality_data_{timestamp}.xlsx")
        
        data.to_csv(csv_file, index=False)
        data.to_excel(excel_file, index=False)
        
        # 2. Create PowerBI report template
        template_file = os.path.join(self.output_dir, f"powerbi_report_template_{timestamp}.txt")
        template_content = self._create_comprehensive_template(data)
        
        with open(template_file, 'w', encoding='utf-8') as f:
            f.write(template_content)
        
        # 3. Create DAX measures file
        dax_file = os.path.join(self.output_dir, f"dax_measures_{timestamp}.txt")
        dax_content = self._create_dax_measures()
        
        with open(dax_file, 'w', encoding='utf-8') as f:
            f.write(dax_content)
        
        # 4. Create PowerBI design guide
        design_guide = os.path.join(self.output_dir, f"powerbi_design_guide_{timestamp}.md")
        design_content = self._create_design_guide(data)
        
        with open(design_guide, 'w', encoding='utf-8') as f:
            f.write(design_content)
        
        # 5. Create sample queries
        queries_file = os.path.join(self.output_dir, f"sample_queries_{timestamp}.sql")
        queries_content = self._create_sample_queries()
        
        with open(queries_file, 'w', encoding='utf-8') as f:
            f.write(queries_content)
        
        return {
            "csv_data": csv_file,
            "excel_data": excel_file,
            "template": template_file,
            "dax_measures": dax_file,
            "design_guide": design_guide,
            "sample_queries": queries_file
        }
    
    def _create_comprehensive_template(self, data: pd.DataFrame) -> str:
        """Create comprehensive PowerBI template"""
        return f"""
# PowerBI Report Template - Honey Quality Analysis
# Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Quick Start

### Step 1: Import Data
1. Open PowerBI Desktop
2. Click "Get Data" → "Text/CSV" or "Excel"
3. Select: honey_quality_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv
4. Click "Load"

### Step 2: Create Charts
- Quality Distribution: quality_category → Bar Chart
- Lab Performance: lab_id + quality_score → Bar Chart
- Time Trends: collection_date + quality_score → Line Chart

## Data Summary
- Total Samples: {len(data):,}
- Average Quality: {data['quality_score'].mean():.2f}
- Premium Rate: {len(data[data['quality_category']=='Premium'])/len(data)*100:.1f}%
- Compliance Rate: {len(data[data['compliance_status']=='Compliant'])/len(data)*100:.1f}%

## Basic Charts
- KPI Cards: Total Samples, Average Quality, Premium Rate
- Quality Distribution: Pie Chart
- Lab Performance: Bar Chart
- Time Trends: Line Chart
        """
    
    def _create_dax_measures(self) -> str:
        """Create comprehensive DAX measures"""
        return """
# PowerBI DAX Measures

## Basic Measures
```
Total Samples = COUNTROWS(honey_quality_data)
Average Quality = AVERAGE(honey_quality_data[quality_score])
Premium Rate = DIVIDE(COUNTROWS(FILTER(honey_quality_data, honey_quality_data[quality_category]="Premium")), COUNTROWS(honey_quality_data), 0)
Compliance Rate = DIVIDE(COUNTROWS(FILTER(honey_quality_data, honey_quality_data[compliance_status]="Compliant")), COUNTROWS(honey_quality_data), 0)
```

## Additional Measures
```
Lab Performance = CALCULATE(AVERAGE(honey_quality_data[quality_score]), ALLEXCEPT(honey_quality_data, honey_quality_data[lab_id]))
Monthly Quality = CALCULATE(AVERAGE(honey_quality_data[quality_score]), DATESMTD(honey_quality_data[collection_date]))
```
        """
    
    def _create_design_guide(self, data: pd.DataFrame) -> str:
        """Create PowerBI design guide"""
        return f"""
# PowerBI Design Guide

## Color Scheme
- Primary: #366092 (Blue)
- Secondary: #70AD47 (Green)
- Accent: #FFC000 (Yellow)

## Layout
- KPI Cards: 200x150 pixels
- Charts: 400x300 pixels
- Title: Segoe UI, 24pt, Bold
- Labels: Segoe UI, 12pt, Regular

## Basic Setup
- Enable cross-filtering between charts
- Use consistent colors
- Keep layout simple and clean
        """
    
    def _create_sample_queries(self) -> str:
        """Create sample SQL queries"""
        return """
# Sample SQL Queries

## Basic Queries
```sql
-- Quality by Lab
SELECT lab_id, AVG(quality_score) as avg_quality, COUNT(*) as samples
FROM honey_quality_data GROUP BY lab_id ORDER BY avg_quality DESC;

-- Quality Trends
SELECT DATE(collection_date) as date, AVG(quality_score) as avg_quality
FROM honey_quality_data GROUP BY DATE(collection_date) ORDER BY date;

-- Compliance Status
SELECT compliance_status, COUNT(*) as count
FROM honey_quality_data GROUP BY compliance_status;
```

## Export for PowerBI
```sql
SELECT * FROM honey_quality_data 
WHERE collection_date >= CURRENT_DATE - INTERVAL '30 days'
ORDER BY collection_date DESC;
```
        """
    
    def generate_powerbi_files(self, data: pd.DataFrame) -> str:
        """Generate PowerBI-ready files"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Export to multiple formats
        csv_file = os.path.join(self.output_dir, f"honey_quality_data_{timestamp}.csv")
        excel_file = os.path.join(self.output_dir, f"honey_quality_data_{timestamp}.xlsx")
        
        data.to_csv(csv_file, index=False)
        data.to_excel(excel_file, index=False)
        
        # Create PowerBI template file
        template_file = os.path.join(self.output_dir, f"powerbi_template_{timestamp}.txt")
        template_content = self._create_powerbi_template(data)
        
        with open(template_file, 'w', encoding='utf-8') as f:
            f.write(template_content)
        
        print(f"PowerBI files generated in: {self.output_dir}")
        return self.output_dir
    
    def _create_powerbi_template(self, data: pd.DataFrame) -> str:
        """Create PowerBI template content"""
        return f"""
# PowerBI Report Template - Honey Quality Analysis
# Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Data Source Files
- CSV: honey_quality_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv
- Excel: honey_quality_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx

## Quick Start Instructions
1. Open PowerBI Desktop
2. Click "Get Data" → "Text/CSV" or "Excel"
3. Select the generated data file
4. Click "Load"
5. Create visualizations as described below

## Recommended Visualizations

### 1. KPI Cards
- Total Samples: {len(data):,}
- Average Quality Score: {data['quality_score'].mean():.2f}
- Premium Quality Rate: {len(data[data['quality_category']=='Premium'])/len(data)*100:.1f}%
- Compliance Rate: {len(data[data['compliance_status']=='Compliant'])/len(data)*100:.1f}%

### 2. Charts
- **Bar Chart**: Quality category distribution
- **Bar Chart**: Laboratory performance comparison
- **Line Chart**: Quality trends over time
- **Scatter Plot**: Moisture vs pH correlation
- **Pie Chart**: Compliance status distribution

## DAX Measures
```
Total Samples = COUNTROWS(honey_quality_data)
Average Quality = AVERAGE(honey_quality_data[quality_score])
Premium Rate = DIVIDE(
    COUNTROWS(FILTER(honey_quality_data, honey_quality_data[quality_category]="Premium")),
    COUNTROWS(honey_quality_data),
    0
)
```

## Data Refresh
- Set up automatic refresh every 6 hours
- Use incremental refresh for large datasets
        """
    
    def generate_all_reports_from_data(self, data: pd.DataFrame) -> dict:
        """Generate all types of reports from existing data"""
        print("Starting automated report generation from existing data...")
        
        try:
            # Generate Excel report
            print("Generating Excel report...")
            excel_report = self.generate_excel_report(data)
            
            # Generate PowerBI files
            print("Generating PowerBI files...")
            powerbi_files = self.generate_powerbi_files(data)
            
            # Generate summary
            summary = {
                "excel_report": excel_report,
                "powerbi_files": powerbi_files,
                "data_summary": {
                    "total_samples": len(data),
                    "avg_quality": data['quality_score'].mean() if 'quality_score' in data.columns else 0,
                    "premium_rate": len(data[data['quality_category']=='Premium']) / len(data) * 100 if 'quality_category' in data.columns else 0,
                    "compliance_rate": len(data[data['compliance_status']=='Compliant']) / len(data) * 100 if 'compliance_status' in data.columns else 0
                }
            }
            
            print("All reports generated successfully from existing data!")
            return summary
            
        except Exception as e:
            print(f"Report generation from existing data failed: {e}")
            raise
    
    def generate_all_reports(self, num_records: int = 1000) -> dict:
        """Generate all types of reports"""
        print("Starting automated report generation...")
        
        # Generate sample data
        print("Generating sample data...")
        data = self.generate_sample_data(num_records)
        
        # Generate Excel report
        print("Generating Excel report...")
        excel_report = self.generate_excel_report(data)
        
        # Generate PowerBI files
        print("Generating PowerBI files...")
        powerbi_files = self.generate_powerbi_files(data)
        
        # Generate summary
        summary = {
            "excel_report": excel_report,
            "powerbi_files": powerbi_files,
            "data_summary": {
                "total_samples": len(data),
                "avg_quality": data['quality_score'].mean(),
                "premium_rate": len(data[data['quality_category']=='Premium']) / len(data) * 100,
                "compliance_rate": len(data[data['compliance_status']=='Compliant']) / len(data) * 100
            }
        }
        
        print("All reports generated successfully!")
        return summary
    
    def generate_complete_powerbi_solution(self, num_records: int = 1000) -> dict:
        """Generate complete PowerBI solution"""
        print("Generating complete PowerBI solution...")
        
        # Generate sample data
        print("Generating sample data...")
        data = self.generate_sample_data(num_records)
        
        # Create all template files
        print("Creating PowerBI template files...")
        files = self.create_powerbi_template_files(data)
        
        # Create summary
        summary = {
            "files": files,
            "data_summary": {
                "total_samples": len(data),
                "avg_quality": data['quality_score'].mean(),
                "premium_rate": len(data[data['quality_category']=='Premium']) / len(data) * 100,
                "compliance_rate": len(data[data['compliance_status']=='Compliant']) / len(data) * 100
            }
        }
        
        print("Complete PowerBI solution generated!")
        return summary


def main():
    """Main function to generate all reports"""
    generator = UnifiedPowerBIGenerator()
    
    try:
        # Generate all reports
        results = generator.generate_all_reports(1000)
        
        print("\nReport Generation Complete!")
        print("=" * 50)
        print(f"Excel Report: {results['excel_report']}")
        print(f"PowerBI Files: {results['powerbi_files']}")
        print("\nData Summary:")
        summary = results['data_summary']
        print(f"   Total Samples: {summary['total_samples']:,}")
        print(f"   Average Quality: {summary['avg_quality']:.2f}")
        print(f"   Premium Rate: {summary['premium_rate']:.1f}%")
        print(f"   Compliance Rate: {summary['compliance_rate']:.1f}%")
        
        print("\nNext Steps:")
        print("1. Open Excel report for detailed analysis")
        print("2. Import PowerBI files to create professional reports")
        
    except Exception as e:
        print(f"Report generation failed: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
